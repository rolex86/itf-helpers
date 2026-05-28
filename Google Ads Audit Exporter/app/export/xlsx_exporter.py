from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.google_ads.report_definitions import REPORT_ORDER, get_report_definition


PERCENT_COLUMNS = {
    "ctr",
    "conversion_rate",
    "search_impression_share",
    "search_budget_lost_impression_share",
    "search_rank_lost_impression_share",
    "target_roas",
    "optimization_score_uplift",
    "engagement_rate",
    "purchase_rate_from_view_item",
}

CURRENCY_COLUMNS = {
    "average_cpc",
    "cost_per_conversion",
    "conversions_value",
    "value_per_conversion",
    "all_conversions_value",
    "total_revenue",
}

DECIMAL_COLUMNS = {
    "conversions",
    "all_conversions",
    "optimization_score",
    "roas",
    "average_session_duration",
    "position",
    "performance_score",
    "accessibility_score",
    "seo_score",
    "best_practices_score",
    "lcp",
    "cls",
    "inp",
    "fcp",
    "speed_index",
}

_MAX_SHEET_NAME_LENGTH = 31
_INVALID_SHEET_NAME_CHARS = re.compile(r"[\[\]:*?/\\]")


def _summary_frame(summary_rows: list[dict[str, Any]]) -> pd.DataFrame:
    return pd.DataFrame(summary_rows, columns=["section", "name", "value", "details", "status", "rows"])


def _ordered_sheets(
    datasets: dict[str, pd.DataFrame],
    flags_df: pd.DataFrame | None,
    derived_sheets: list[tuple[str, pd.DataFrame]],
) -> list[tuple[str, pd.DataFrame]]:
    ordered: list[tuple[str, pd.DataFrame]] = []
    if flags_df is not None:
        ordered.append(("Basic flags", flags_df))
    ordered.extend(derived_sheets)
    for report_key in REPORT_ORDER:
        if report_key in datasets:
            ordered.append((get_report_definition(report_key).sheet_name, datasets[report_key]))
    for report_key, dataframe in datasets.items():
        if report_key in REPORT_ORDER:
            continue
        ordered.append((report_key, dataframe))
    return ordered


def _clean_sheet_name(sheet_name: str) -> str:
    cleaned = _INVALID_SHEET_NAME_CHARS.sub("-", str(sheet_name or "")).strip()
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:_MAX_SHEET_NAME_LENGTH]


def _unique_sheet_name(sheet_name: str, used_names: set[str]) -> str:
    base = _clean_sheet_name(sheet_name)
    if base not in used_names:
        used_names.add(base)
        return base

    counter = 2
    while True:
        suffix = f" {counter}"
        candidate = f"{base[: _MAX_SHEET_NAME_LENGTH - len(suffix)]}{suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        counter += 1


def _apply_common_formatting(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for column_cells in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        width = max((len(value) for value in values), default=0) + 2
        width = min(max(width, 12), 60)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _apply_number_formats(worksheet) -> None:
    headers = [worksheet.cell(row=1, column=index).value for index in range(1, worksheet.max_column + 1)]
    for col_index, header in enumerate(headers, start=1):
        if not header:
            continue
        number_format = None
        if header in PERCENT_COLUMNS:
            number_format = "0.00%"
        elif header in CURRENCY_COLUMNS:
            number_format = '#,##0.00 [$-en-US]$'
        elif header.endswith("_micros"):
            number_format = "#,##0"
        elif header in DECIMAL_COLUMNS:
            number_format = "0.00"

        if number_format is None:
            continue

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=col_index).number_format = number_format


def export_workbook(
    xlsx_path: Path,
    summary_rows: list[dict[str, Any]],
    datasets: dict[str, pd.DataFrame],
    flags_df: pd.DataFrame | None,
    derived_sheets: list[tuple[str, pd.DataFrame]] | None = None,
) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    summary_df = _summary_frame(summary_rows)
    derived_sheets = derived_sheets or []
    used_sheet_names: set[str] = set()

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary_sheet_name = _unique_sheet_name("Summary", used_sheet_names)
        summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)

        for raw_sheet_name, dataframe in _ordered_sheets(
            datasets=datasets,
            flags_df=flags_df,
            derived_sheets=derived_sheets,
        ):
            sheet_name = _unique_sheet_name(raw_sheet_name, used_sheet_names)
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            _apply_common_formatting(worksheet)
            _apply_number_formats(worksheet)
